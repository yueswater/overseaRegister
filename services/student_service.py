import os
from openpyxl import Workbook, load_workbook
from db.models.student import Student
from db.models.payment import Payment
from db.models.student_record import StudentRecord
from typing import Dict

REGISTERED_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "registered_students.xlsx")

def from_form(form: Dict) -> StudentRecord:
    student_data = {
        "series_id": form.get("series_id", ""),
        "name": form.get("name", ""),
        "is_pass_exam": form.get("is_pass_exam", "false") in ("true", "1", "on"),
        "class_id": form.get("class_id", ""),
        "student_id": form.get("student_id", "")
    }

    payment_data = {
        key[len("payment["):-1]: int(value)
        for key, value in form.items()
        if key.startswith("payment[")
    }

    student = Student.from_dict(student_data)
    payment = Payment.from_dict(payment_data)
    return StudentRecord(student=student, payment=payment)


def save_student_record(record: StudentRecord):
    # Create data folder
    os.makedirs(os.path.dirname(REGISTERED_FILE), exist_ok=True)

    # column content
    row = [
        record.student.series_id,
        record.student.name,
        record.student.class_id,
        record.student.student_id,
        record.student.is_pass_exam,
    ] + [record.payment.items.get(key, 0) for key in sorted(record.payment.items.keys())]

    if os.path.exists(REGISTERED_FILE):
        wb = load_workbook(REGISTERED_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Add the header
        header = [
            "流水號", "姓名", "班級", "學號", "是否通過測驗"
        ] + sorted(record.payment.items.keys())
        ws.append(header)

    ws.append(row)
    wb.save(REGISTERED_FILE)